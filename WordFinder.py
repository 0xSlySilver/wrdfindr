from pathlib import Path 
import typer
import csv
from openpyxl import load_workbook
import json
import xml.etree.ElementTree as ET
import pdfplumber
from docx import Document
import configparser
import yaml
from io import StringIO
import logging


logging.getLogger("pdfminer").setLevel(logging.ERROR)
app = typer.Typer()

 # Typer command flags
@app.command(help="Search files in a directory for a given word.")
def main(
    keyword: str = typer.Option(..., "-w", "--word", help="Word to search for"),
    directory: Path = typer.Option(..., "-d", "--directory", help="Directory to search"),
    extensions: str = typer.Option("*", "-e", help="Comma-separated file extensions to search. Ex: -e pdf,txt"),
    case_sensitive: bool = typer.Option(
        False,
        "-c",
        "--case-sensitive",
        help="Enable case sensitive search"
    ),
    recursive: bool = typer.Option(
         False,
         "-r",
         "--recursive",
         help="Recursively scans the directory."
    ),
    output: Path | None = typer.Option(
        None,
        "-o",
        "--output", 
        help="csv file to output report to"),

):
    if not directory.is_dir():
        typer.echo("Invalid directory")
        raise typer.Exit(code=1)

    # Normalizes extensions
    extensions_list = [e.strip().lower() for e in extensions.split(",")]
    
    typer.echo(
        f"Searching for '{keyword}' in {directory} "
        f"for extensions: {extensions_list}"
    )
        
    total, results = search_directory(
        keyword,
        directory,
        extensions_list,
        case_sensitive,
        recursive,
    )

    typer.echo(f"Word found {total} times.")

    if output:
        write_csv(output, results)
        typer.echo(f"Results written to {output} in current folder.")


# ------- Extractor Logic ------- #
def read_txt(path: Path) -> str:
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def read_pdf(path: Path) -> str:
    text = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text
    return text

def read_docx(path: Path) -> str:
    doc = Document(path)
    text = []
    for para in doc.paragraphs:
        text.append(para.text)
    return str(text)

def read_xlsx(path: Path) -> str:
    workbook = load_workbook(path)
    text = []
    for sheet in workbook:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    text.append(str(cell))
    return str(text)

def read_csv(path: Path) -> str:
    text = []
    with open(path, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            if row is not None:
               text.append(row)
    return str(text)

def read_json(path: Path) -> str:
    with open(path, 'r') as f:
        data = json.load(f)
        json_string = json.dumps(data)
    return json_string

def read_xml(path: Path) -> str:
    tree = ET.parse(path) 
    root = tree.getroot()
    text = ET.tostring(root, encoding='unicode')
    return text

def read_yaml(path: Path) -> str:
    with open (path, 'r') as f:
        text = yaml.safe_load(path.read_text())
        return str(text)

def read_ini(path:Path) -> str:
    config = configparser.ConfigParser()
    config.read(path)
    string_buffer = StringIO()
    config.write(string_buffer)
    text = string_buffer.getvalue()
    return str(text)

EXTRACTORS = {
    "txt": read_txt,
    "md": read_txt,
    "log": read_txt,
    "env": read_txt,
    "pdf": read_pdf,
    "docx": read_docx,
    "json": read_json,
    "xml": read_xml,
    "yaml": read_yaml,
    "yml": read_yaml,
    "xlsx": read_xlsx,
    "csv": read_csv,
    "ini": read_ini,
}

# Directory iteration logic
def search_directory(
    keyword: str,
    directory: Path,
    extensions: list[str],
    case_sensitive: bool,
    recursive: bool,
) -> tuple[int, list[dict]]:

    total_words = 0
    results: list[dict] = []

    iterator = directory.rglob("*") if recursive else directory.iterdir()

    for path in iterator:
        if not path.is_file():
            continue

        # Removes the . from files    
        ext = path.suffix.lstrip(".").lower()

        # Filters extensions
        if extensions != ["*"] and ext not in extensions:
            continue

        extractor = EXTRACTORS.get(ext)
        # Continues if extractor is unsupported
        if not extractor:
            continue  

        try:
            content = extractor(path)
            
            if not case_sensitive:
                content = content.lower()
                keyword = keyword.lower()
            
            matches = content.count(keyword)

            if matches > 0:
                # File contents
                typer.echo(f"Found in: {path} | {matches} occurence(s)")
                results.append({
                    "file": str(path),
                    "extension": ext,
                    "count": matches,
    })

            total_words += matches
            
        except Exception as e:
            print(f"Error reading file {path}: {e}")

    
    
    return total_words, results

# Writes to csv for output
def write_csv(output: Path, results: list[dict]):
    with output.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["file", "extension", "count"]
        )
        writer.writeheader()
        writer.writerows(results)
       
if __name__ == "__main__":
    app()

